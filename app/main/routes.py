# -*- coding: utf-8 -*-
from flask import render_template, redirect, url_for, request
from app import db
from app.main.forms import FilterForm
from flask_login import current_user, login_required
from app.models import User, Devinfo, Department
import openpyxl
from datetime import datetime
from app.main import bp
from app.usbparam import status_types


@bp.route('/', methods=['GET', 'POST'])
@bp.route('/index', methods=['GET', 'POST'])
@login_required
def index():
    filterform = FilterForm()
    filterform.department.choices = filterform.department.choices + [(dep.id, dep.name) for dep in Department.query.order_by('name').all()]
    devices = Devinfo.query.filter_by(last_state=True).order_by(Devinfo.rec_date.desc()).all()
    if filterform.validate_on_submit():
        if filterform.submit_apply.data:
            if filterform.dev_numb.data != "":
                devices = [dev for dev in devices if filterform.dev_numb.data in dev.dev_numb]
            if filterform.dev_id.data != "":
                devices = [dev for dev in devices if filterform.dev_id.data in dev.dev_id]
            if filterform.department.data != "None":
                devices = [dev for dev in devices if dev.department_id == int(filterform.department.data)]
            if filterform.doc_numb.data != "":
                devices = [dev for dev in devices if filterform.doc_numb.data in dev.doc_numb]
            if filterform.status.data != "None":
                devices = [dev for dev in devices if dev.status == filterform.status.data]
            if filterform.owner.data != "":
                devices = [dev for dev in devices if filterform.owner.data in dev.owner]
            elif filterform.sortradio.data == "dev_numb":
                devices = sorted(devices, key=lambda dev: dev.dev_numb)
            elif filterform.sortradio.data == "department_name":
                devices = sorted(devices, key=lambda dev: dev.department.name)
        elif filterform.submit_clear.data:
            return redirect(url_for('main.index'))

    return render_template('index.html', title='Список зарегистрированных USB-устройств',
                           devices=devices, form=filterform)


@bp.route('/load_from_file', methods=['POST', 'GET'])
def load_from_file():
    if request.method == "POST":
        # do something
        path_from = request.form['path_from']
        mytable = path_from.maketrans('', '', ' \n\t\r')
        path_from = path_from.translate(mytable)
        if path_from == "":
            return render_template("load_from_file.html", wrong_path=True)

        # Define variable to load the wookbook
        wookbook = openpyxl.load_workbook(path_from)
        # Define variable to read the active sheet:
        worksheet = wookbook.active
        for line in worksheet.iter_rows(2, worksheet.max_row):
            department_id = get_dep(line[1].value).id
            dev_id = line[2].value

            list_owner_numb = line[3].value.split("/")
            owner = list_owner_numb[0]
            dev_numb = line[1].value.split('_')[1] + '-' + list_owner_numb[1]

            doc_numb = line[4].value

            if "_от" in line[4].value:
                rec_date = line[4].value.split("_от")[1].replace("г.", "").replace(".", "-")
                format = '%d-%m-%y'
                rec_date = datetime.strptime(rec_date, format)
            else:
                rec_date = datetime.now()

            remark = ""
            if line[5].value != None:
                remark = line[5].value

            if "вышла из строя" in remark or "сдана" in remark:
                # добавляем неактивную запись флешки и активную с новой датой
                devinfo = Devinfo(dev_numb=dev_numb, dev_id=dev_id, department_id=department_id,
                                  owner=owner, doc_numb=doc_numb, doc_ref="", rec_date=rec_date, remark=remark,
                                  last_state=False)
                try:
                    db.session.add(devinfo)
                    db.session.commit()
                except:
                    return "При загрузке из файла произошла ошибка. Строка в файле: " + line[0].value

                if "/" in remark:
                    remark = remark.split("/")[0]

                status = ""
                if "вышла из строя" in remark:
                    status = "Вышла из строя"
                elif "сдана" in remark:
                    status = "Сдана"

                rec_date = remark.split("_")[1].replace('.', '-')
                format = '%d-%m-%Y'
                rec_date = datetime.strptime(rec_date, format)

                # print('0', line[0].value, ' 1', line[1].value, ' 2', line[2].value, ' 3', line[3].value, ' 4',
                #      line[4].value, ' 5', line[5].value)
                # print(line[0].value, ' ', department_id, " ", dev_id, " ", owner, ' ', dev_numb, ' ', doc_numb, ' ',
                #      rec_date, status)
                # print('\n')

                devinfo = Devinfo(dev_numb=dev_numb, dev_id=dev_id, department_id=department_id,
                                  owner=owner, doc_numb=doc_numb, doc_ref="", status=status, rec_date=rec_date,
                                  remark=remark)
                try:
                    db.session.add(devinfo)
                    db.session.commit()
                except:
                    return " 1 При загрузке из файла произошла ошибка. Строка в файле: " + line[0].value

            else:
                # добавляем только активную запись
                devinfo = Devinfo(dev_numb=dev_numb, dev_id=dev_id, department_id=department_id,
                                  owner=owner, doc_numb=doc_numb, doc_ref="", rec_date=rec_date, remark=remark)
                try:
                    db.session.add(devinfo)
                    db.session.commit()
                except:
                    return "2 При загрузке из файла произошла ошибка. Строка в файле: " + str(line[0].value)

        return redirect(url_for('main.index'))
    else:
        return render_template("load_from_file.html", wrong_path=False)


def get_dep(department_name):
    department = Department.query.filter_by(name=department_name).first()

    if department is None:
        dep = Department(name=department_name)
        try:
            db.session.add(dep)
            db.session.commit()
            return dep
        except:
            return None
    return department

