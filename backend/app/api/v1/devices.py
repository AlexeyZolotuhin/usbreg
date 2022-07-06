from app.api.v1 import bp
from app.api.v1.errors import bad_request
from app.api.v1.auth import token_auth
from app.models import Devinfo, Department
import json
from app import db
from flask import request
import datetime
import openpyxl


@bp.route('/devices/all', methods=["GET"])
@token_auth.login_required
def get_devices_all():
    response_departments = {dep.id: dep.name for dep in Department.query.order_by('name').all()}
    devices = Devinfo.query.filter_by(last_state=True).order_by(Devinfo.rec_date.desc()).all()
    response_devices = [dev.to_dict() for dev in devices]
    response = {'departments': response_departments, 'devices': response_devices}
    return json.dumps(response, ensure_ascii=False).encode('utf8')


@bp.route('/devices/all/with_filter', methods=["GET"])
@token_auth.login_required
def get_filtered_devices():
    devices = Devinfo.query.filter_by(last_state=True).order_by(Devinfo.rec_date.desc()).all()
    data = request.get_json(force=True) or {}

    if 'department_id' in data:
        data['department_id'] = int(data['department_id'])

    fields = ['dev_numb',  'department_id', 'dev_id', 'doc_numb', 'status', 'owner']
    for field in fields:
        if field in data:
            if field == 'department_id':
                devices = [dev for dev in devices if data[field] == dev.get_value_field(field)]
                continue
            devices = [dev for dev in devices if data[field] in dev.get_value_field(field)]
    response_devices = sorted([dev.to_dict() for dev in devices],
                              key=lambda dev: dev[data['sortradio']])

    return json.dumps(response_devices, ensure_ascii=False).encode('utf8')


@bp.route('/devices/<int:id>', methods=["GET"])
@token_auth.login_required
def get_device_rec(id):
    return json.dumps(Devinfo.query.get_or_404(id).to_dict(),
                      ensure_ascii=False).encode('utf8')


@bp.route('/devices/<int:id>', methods=["PUT"])
@token_auth.login_required
def update_rec_dev(id):
    rec_dev = Devinfo.query.get_or_404(id)
    data = request.get_json(force=True) or {}
    if 'dev_numb' in data and \
            Devinfo.query.filter_by(dev_numb=data['dev_numb'], status='Активная', last_state=True).first():
        return bad_request('Активное устройство с таким именем уже существует.')

    if 'department_id' in data:
        data['department_id'] = int(data['department_id'])
    try:
        rec_dev.from_dict(data)
        db.session.commit()
    except:
        return bad_request("При обновлении записи об устройстве произошла ошибка")
    return json.dumps(rec_dev.to_dict(), ensure_ascii=False).encode('utf8')


@bp.route('/devices/detail', methods=["GET"])
@token_auth.login_required
def get_device_detail():
    data = request.get_json(force=True) or {}
    detail_dev = Devinfo.query.filter_by(dev_id=data['dev_id']).order_by(Devinfo.rec_date.desc()).all()
    detail_dev = [dev.to_dict() for dev in detail_dev]

    if 'dev_id' in data:
        return json.dumps(detail_dev)
    return bad_request('No data')


@bp.route('/devices/del_all_data/<string:url_dev_id>', methods=["DELETE"])
@token_auth.login_required
def device_delete_all(url_dev_id):
    dev_id = url_dev_id.replace('*', '\\')
    del_dev = Devinfo.query.filter_by(dev_id=dev_id).all()
    try:
        for row in del_dev:
            db.session.delete(row)
        db.session.commit()
    except:
        return bad_request("При удалении всех записей об устройстве произошла ошибка")

    return '', 204


@bp.route('/devices/del_rec/<int:id>', methods=["DELETE"])
@token_auth.login_required
def delete_device_rec(id):
    del_rec = Devinfo.query.get_or_404(id)
    try:
        db.session.delete(del_rec)
        db.session.commit()
        if del_rec.last_state:
            dev_detail = Devinfo.query.filter_by(dev_id=del_rec.dev_id).order_by(Devinfo.rec_date).all()
            if dev_detail is not None:
                dev_detail[-1].last_state = True
                db.session.commit()
    except:
        return bad_request("При удалении записи об устройстве произошла ошибка")

    return '', 204


@bp.route('/devices/new_rec_dev', methods=['POST'])
@token_auth.login_required
def add_new_rec_dev():
    data = request.get_json(force=True)
    rec_dev = Devinfo.query.get_or_404(int(data['id']))
    if 'department_id' in data:
        data['department_id'] = int(data['department_id'])
    if 'dev_numb' in data and data['dev_numb'] != rec_dev.dev_numb and\
            Devinfo.query.filter_by(dev_numb=data['dev_numb'], status='Активная', last_state=True).first():
        return bad_request('Активное устройство с таким именем уже существует.')

    edit_dev = Devinfo()
    try:
        edit_dev.from_dict(data)
        rec_dev.last_state = False
        db.session.add(edit_dev)
        db.session.commit()
    except:
        return bad_request("При изменении данных об устройстве произошла ошибка")
    response = json.dumps(edit_dev.to_dict(), ensure_ascii=False).encode('utf8')
    return response


@bp.route('/devices/add_dev', methods=['POST'])
@token_auth.login_required
def add_new_dev():
    data = request.get_json(force=True)
    data['department_id'] = int(data['department_id'])
    if Devinfo.query.filter_by(dev_numb=data['dev_numb'], status='Активная', last_state=True).first():
        return bad_request('Активное устройство с таким именем уже существует.')
    if Devinfo.query.filter_by(dev_id=data['dev_id']).first():
        return bad_request('Устройство с указанным идентификатором уже существует.')
    new_dev = Devinfo()
    try:
        new_dev.from_dict(data)
        db.session.add(new_dev)
        db.session.commit()
    except:
        return bad_request("При добавдение нового устройства произошла ошибка")
    response = json.dumps(new_dev.to_dict(), ensure_ascii=False).encode('utf8')
    return response


@bp.route('/devices/load_from_file', methods=['POST'])
@token_auth.login_required
def load_from_file():
    filename = request.get_json(force=True)
    data = get_devices_from_file(filename['filename'])
    report = []
    if data is None:
        report.append({'N': 0, 'dev_numb': 'none', 'result': 'Ошибка',
                       'message': 'Возникла ошибка во время парсинга файла. Проверте правильность заполнения файла'})
    else:
        for i, dev in enumerate(data):
            if dev['status'] == 'Активная' and dev['last_state'] is True:
                if Devinfo.query.filter_by(dev_numb=dev['dev_numb'], status='Активная', last_state=True).first() and \
                'na' not in dev['dev_numb']:
                     report.append({'N': i, 'dev_numb': dev['dev_numb'], 'result': 'Ошибка',
                                    'message': 'активное устройство с таким именем уже существует'})
                     continue

                if Devinfo.query.filter_by(dev_id=dev['dev_id'], status='Активная', last_state=True).first():
                    report.append({'N': i, 'dev_numb': dev['dev_numb'], 'result': 'Ошибка',
                                   'message': 'активное устройство с указанным идентификатором уже существует'})
                    continue
            try:
                if dev['rec_date'] is not None:
                    format_date = '%d-%m-%y'
                    dev['rec_date'] = datetime.datetime.strptime(dev['rec_date'], format_date)
            except:
                report.append({'N': i, 'dev_numb': dev['dev_numb'], 'result': 'Ошибка',
                               'message': 'ошибка формата даты'})
                continue

            try:
                dev['department_id'] = Department.query.filter_by(name=dev['department_name']).first().id
            except:
                report.append({'N': i, 'dev_numb': dev['dev_numb'], 'result': 'Ошибка',
                               'message': f'указанного подразделения в БД не найдено: {dev["department_name"]}'})
                continue

            new_dev = Devinfo()
            try:
                new_dev.from_dict(dev)
                db.session.add(new_dev)
                db.session.commit()
                report.append({'N': i, 'dev_numb': dev['dev_numb'], 'result': 'Успешно',
                               'message': 'запись в БД об устройстве добавлена успешно'})
            except:
                report.append({'N': i, 'dev_numb': dev['dev_numb'], 'result': 'Ошибка',
                               'message': 'при добавлении данных в БД произошла ошибка'})

    response = json.dumps(report, ensure_ascii=False).encode('utf8')
    return response


def get_devices_from_file(filename):
    data = []
    try:
        # Define variable to load the wookbook
        wook_book = openpyxl.load_workbook(filename)
        # Define variable to read the active sheet:
        worksheet = wook_book.active
        for line in worksheet.iter_rows(2, worksheet.max_row):
            department_name = line[1].value
            dev_id = line[2].value

            list_owner_numb = line[3].value.split("/")
            owner = list_owner_numb[0]
            dev_numb = line[1].value.split('_')[1] + '-' + list_owner_numb[1]

            doc_numb = line[4].value

            if "_от" in line[4].value:
                rec_date = line[4].value.split("_от")[1].replace("г.", "").replace(".", "-")
            else:
                rec_date = None

            remark = ""
            if line[5].value is not None:
                remark = line[5].value

            if "вышла из строя" in remark or "сдана" in remark:
                # добавляем неактивную запись флешки и активную с новой датой
                data.append({'dev_numb': dev_numb, 'dev_id': dev_id, 'department_name': department_name,
                             'owner': owner, 'doc_numb': doc_numb, 'rec_date': rec_date, 'remark': remark,
                             'status': 'Активная', 'last_state': False})

                remark1 = None
                if "/" in remark:
                    remark, remark1 = remark.split("/")

                status = ""
                if "вышла из строя" in remark:
                    status = "Вышла из строя"
                elif "сдана" in remark:
                    status = "Сдана"

                rec_date = remark.split("_")[1].replace('.', '-')
                data.append({'dev_numb': dev_numb, 'dev_id': dev_id, 'department_name': department_name,
                             'owner': owner, 'doc_numb': doc_numb, 'status': status, 'rec_date': rec_date,
                             'remark': remark if remark1 is None else remark1, 'last_state': True})
            else:
                # добавляем только активную запись
                data.append({'dev_numb': dev_numb, 'dev_id': dev_id, 'department_name': department_name,
                             'owner': owner, 'doc_numb': doc_numb, 'status': "Активная", 'rec_date': rec_date,
                             'remark': remark, 'last_state': True})
    except:
        data = None

    return data
